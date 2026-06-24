# -*- coding: utf-8 -*-
"""
Actualizar Dashboard AMSA — Actividad Corporativa + Distribuibles
------------------------------------------------------------------
Herramienta con interfaz gráfica para regenerar los datos del dashboard
a partir de los Excel y reinyectarlos en el HTML autocontenido.

Requisitos: Python 3 + openpyxl  (pip install openpyxl).  NO requiere Excel.

Uso: doble clic en "Actualizar Dashboard.bat" (o ejecutar este .py).
  1) Elige el Excel corporativo (BBDD ... Consolidado).
  2) Elige el Excel de distribuibles (pestañas MLP/ANT/CEN/CMZ).
  3) Clic en "Actualizar dashboard".
El HTML se actualiza y se genera el .zip listo para enviar.
"""
import os, re, json, gzip, base64, shutil, zipfile, datetime, traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
HTML_NAME = "Dashboard Actividad Corporativa.html"
DATA_UUID = "4220c217-9166-414f-8f44-a89ae6c16113"   # asset data.js dentro del HTML

# Diccionario VP (nombre en distribuibles -> clave VP del corporativo/BBDD)
VP_DICT = {
    "VP Desarrollo": "VP Desarrollo",
    "VP Planificación y Servicios Técnicos": "VP Plani y Serv Tecn",
    "VP Proyectos": "VPP",
    "VP Finanzas": "VP Finanzas",
    "VP Asuntos Corporativos": "VPAC",
    "VP Legal": "VPL",
    "VP Comercialización": "VPC",
    "VP Estrategia e Innovación": "VP Estrategia e Inno",
    "VP Sustentabilidad": "VP Sustentabilidad",
    "VP Personas y Organización": "VP Recursos Humanos",
    "Auditoría (HH Corporativo)": "Gcia Auditoría",
}

# --- Fuente corporativa nueva: "Gastos Act Corpor histórico 2022-2025 (original).xlsx" ---
# Hoja con el máximo detalle: col A = Gerencia (con relleno hacia abajo), col B = Ítem,
# años Real/Plan en cols 2-3 (2022), 5-6 (2023), 8-9 (2024), 11-12 (2025) (0-based).
SHEET_CORP = "Act Corpo VP+Gerencia+items (2)"

# Algunas gerencias vienen truncadas/abreviadas en el export nuevo; se normalizan
# al nombre que usa el dashboard (verificado contra el data.js validado).
GER_RENAME = {
    "Gcia Téc d Min y Pro": "Gerencia Minería y Procesos",
    "Prog Compet Costos": "Programa competitividad de costos",
    "Generalistas AMSA": "Generalistas",
    "Gcia.Pr.nva form tra": "Grcia Proy. Nuevas formas de trabajar",
    "Gcia TICA": "TICA corporativo",
    "Transformación Digit": "Transformación Digital",
}

# Gerencia (nombre dashboard) -> Vicepresidencia.  El nuevo Excel no trae la VP;
# este mapeo reproduce el del dashboard validado (las gerencias del CECOS usan
# abreviaturas distintas, por eso no se puede cruzar por nombre de forma fiable).
GER2VP = {
    "VP Desarrollo": "VP Desarrollo",
    "Gcia de Desarrollo": "VP Desarrollo",
    "Evaluación de Negoci": "VP Desarrollo",
    "Evaluación Recursos": "VP Desarrollo",
    "Servicios Técnicos": "VP Desarrollo",
    "Gncia Gest Ser y Ctr": "VP Desarrollo",
    "Geología Corporativa": "VP Desarrollo",
    "Inteligencia Minera": "VP Desarrollo",
    "MIC Costa": "VP Desarrollo",
    "Gcia de Sustentab": "VP Desarrollo",
    "VP Plani y Serv Tecn": "VP Plani y Serv Tecn",
    "Gerencia Minería": "VP Plani y Serv Tecn",
    "Gerencia Procesos": "VP Plani y Serv Tecn",
    "Gerencia Minería y Procesos": "VP Plani y Serv Tecn",
    "Gerencia Relaves": "VP Plani y Serv Tecn",
    "Gercia Mantenimiento": "VP Plani y Serv Tecn",
    "Chief Operating Offi": "VP Plani y Serv Tecn",
    "Gcia Recursos Minero": "VP Plani y Serv Tecn",
    "Gcia Recurs Hídricos": "VP Plani y Serv Tecn",
    "Gcia Excel Operacion": "VP Plani y Serv Tecn",
    "VPP": "VPP",
    "Soporte AMSA TMM": "VPP",
    "Presidencia Ejecutiv": "Presidencia Ejecutiv",
    "Gcia Auditoría": "Gcia Auditoría",
    "Directorio": "Directorio",
    "VP Finanzas": "VP Finanzas",
    "Gcia Abastecimiento": "VP Finanzas",
    "Gcia Contabilidad": "VP Finanzas",
    "Gcia Planificación": "VP Finanzas",
    "Gcia Inv y Finanzas": "VP Finanzas",
    "Gcia Impuestos": "VP Finanzas",
    "Programa competitividad de costos": "VP Finanzas",
    "Grcia.Riesg.CompyCIn": "VP Finanzas",
    "Gcia Planif Financie": "VP Finanzas",
    "Seguros Corp": "VP Finanzas",
    "Gcia Competitividad": "VP Finanzas",
    "Gcia Eng y Serv Estr": "VP Finanzas",
    "Depto Serv Generales": "VP Finanzas",
    "VP Recursos Humanos": "VP Recursos Humanos",
    "Administ de Personal": "VP Recursos Humanos",
    "G de Relaciones Lab": "VP Recursos Humanos",
    "Compensaciones": "VP Recursos Humanos",
    "D° Organizacional": "VP Recursos Humanos",
    "Gercia Rem y Gestión": "VP Recursos Humanos",
    "Sistemas RH": "VP Recursos Humanos",
    "Calidad de Vida": "VP Recursos Humanos",
    "Admin San Lorenzo": "VP Recursos Humanos",
    "Admin Programa JP": "VP Recursos Humanos",
    "SubGn Remuneraciones": "VP Recursos Humanos",
    "Progr.Diver e Inclus": "VP Recursos Humanos",
    "Efect.Organizacional": "VP Recursos Humanos",
    "Generalistas": "VP Recursos Humanos",
    "Reclutam y Selección": "VP Recursos Humanos",
    "Grcia Proy. Nuevas formas de trabajar": "VP Recursos Humanos",
    "Aprendizaje RH": "VP Recursos Humanos",
    "VPAC": "VPAC",
    "Gcia de Asuntos Públ": "VPAC",
    "Gcia de Comunicaci": "VPAC",
    "Subg Protección Indu": "VPAC",
    "Asuntos Corp. Norte": "VPAC",
    "Asuntos Corp. MLP": "VPAC",
    "VPL": "VPL",
    "Prop Minera": "VPL",
    "Mant Prop Minera": "VPL",
    "Const Prop Minera": "VPL",
    "VPC": "VPC",
    "VP Rel institucional": "VP Rel institucional",
    "Energía": "VP Estrategia e Inno",
    "Innovación": "VP Estrategia e Inno",
    "TICA corporativo": "VP Estrategia e Inno",
    "VP Estrategia e Inno": "VP Estrategia e Inno",
    "Data y Analít Avanza": "VP Estrategia e Inno",
    "Transformación Digital": "VP Estrategia e Inno",
    "Geren de Descarboniz": "VP Estrategia e Inno",
    "Gerencia Rep y Sust": "VP Sustentabilidad",
    "Gcia Medioambiente": "VP Sustentabilidad",
    "Gerencia de S&SO": "VP Sustentabilidad",
    "VP Sustentabilidad": "VP Sustentabilidad",
}


# ----------------------------- utilidades ------------------------------------
def _num(v):
    """Celda -> número redondeado a 2 decimales (entero si es exacto), o None."""
    if v is None or v == "":
        return None
    try:
        f = round(float(v), 2)
    except (ValueError, TypeError):
        return None
    return int(f) if f == int(f) else f


def _txt(v):
    return "" if v is None else str(v).strip()


def _uniq(seq):
    seen, out = set(), []
    for x in seq:
        if x not in seen:
            seen.add(x); out.append(x)
    return out


# --------------------------- lectura de Excel --------------------------------
def leer_bbdd(path, log):
    """Hoja '""" + SHEET_CORP + """': col A = Gerencia (relleno hacia abajo),
    col B = Ítem, años Real/Plan en cols 2-3 / 5-6 / 8-9 / 11-12 (0-based).
    La VP se deriva de la gerencia con GER2VP; nombres truncados se normalizan
    con GER_RENAME."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = next((s for s in wb.sheetnames if s.strip() == SHEET_CORP), None)
    if sheet is None:
        raise ValueError(f"El Excel corporativo no tiene la hoja '{SHEET_CORP}'.")
    ws = wb[sheet]
    records, ger = [], None
    sin_vp = set()
    for row in ws.iter_rows(min_row=8, values_only=True):
        c0, c1 = row[0], row[1]
        if c0 not in (None, ""):
            ger = GER_RENAME.get(_txt(c0), _txt(c0))
        if c1 in (None, ""):
            continue
        vp = GER2VP.get(ger)
        if vp is None:
            sin_vp.add(ger)
            vp = ger
        g = lambda i: row[i] if i < len(row) else None
        records.append({
            "vp": vp, "ger": ger, "item": _txt(c1),
            "y2022": {"real": _num(g(2)),  "plan": _num(g(3))},
            "y2023": {"real": _num(g(5)),  "plan": _num(g(6))},
            "y2024": {"real": _num(g(8)),  "plan": _num(g(9))},
            "y2025": {"real": _num(g(11)), "plan": _num(g(12))},
        })
    wb.close()
    if sin_vp:
        log("  AVISO: gerencias sin VP en GER2VP (revisar): " + ", ".join(sorted(sin_vp)))
    log(f"  Corporativo: {len(records)} registros.")
    return records


def leer_distribuibles(path, log):
    """Pestañas MLP/ANT/CEN/CMZ. Localiza columnas de año por '2022.TOTAL' y los
    encabezados VP / Item Relevante / CECO. VP se mapea a la clave del BBDD."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    comps = ["MLP", "ANT", "CEN", "CMZ"]
    faltan = [c for c in comps if c not in wb.sheetnames]
    if faltan:
        raise ValueError("Al Excel de distribuibles le faltan pestañas: " + ", ".join(faltan))
    dist, sin_mapear = [], set()
    for cm in comps:
        ws = wb[cm]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]  # 0-based
        # fila de años
        hy = None
        for ri in range(min(15, len(grid))):
            if any(str(c) == "2022.TOTAL" for c in grid[ri]):
                hy = ri; break
        if hy is None:
            raise ValueError(f"No encontré los encabezados de año en la pestaña {cm}.")
        yc = {}
        for ci, c in enumerate(grid[hy]):
            for y in (2022, 2023, 2024, 2025):
                if str(c) == f"{y}.TOTAL":
                    yc[y] = ci
        hl = grid[hy + 1]
        vpC = itC = cecoC = None
        for ci, h in enumerate(hl):
            if h == "VP": vpC = ci
            elif h == "Item Relevante": itC = ci
            elif h == "CECO": cecoC = ci
        n = 0
        for ri in range(hy + 2, len(grid)):
            row = grid[ri]
            it = row[itC] if itC < len(row) else None
            if it in (None, ""):
                continue
            vp_raw = _txt(row[vpC]) if vpC < len(row) else ""
            vp = VP_DICT.get(vp_raw, vp_raw)
            if vp_raw not in VP_DICT:
                sin_mapear.add(vp_raw)
            ger = _txt(row[cecoC]) if cecoC < len(row) else ""
            g = lambda i: row[i] if i is not None and i < len(row) else None
            dist.append({
                "comp": cm, "vp": vp, "ger": ger, "item": _txt(it),
                "y2022": {"real": _num(g(yc[2022])), "plan": _num(g(yc[2022] + 1))},
                "y2023": {"real": _num(g(yc[2023])), "plan": _num(g(yc[2023] + 1))},
                "y2024": {"real": _num(g(yc[2024])), "plan": _num(g(yc[2024] + 1))},
                "y2025": {"real": _num(g(yc[2025])), "plan": _num(g(yc[2025] + 1))},
            })
            n += 1
        log(f"  {cm}: {n} registros.")
    wb.close()
    if sin_mapear:
        log("  AVISO: VP sin mapear (se usan tal cual): " + ", ".join(sorted(sin_mapear)))
    log(f"  Distribuibles: {len(dist)} registros.")
    return dist


# ---------------------- leer/escribir el HTML --------------------------------
def _asset_regex():
    return re.compile(
        r'("' + re.escape(DATA_UUID) +
        r'":\{"mime":"[^"]*","compressed":true,"data":")([^"]*)(")')


def maps_actuales(html):
    """Recupera itemNames/vpNames/gerNames del data.js embebido (se conservan)."""
    m = _asset_regex().search(html)
    if not m:
        raise ValueError("No encontré el bloque de datos (data.js) dentro del HTML.")
    data_js = gzip.decompress(base64.b64decode(m.group(2))).decode("utf-8")
    ada_part = data_js.split("window.DIST_DATA", 1)[0]
    ada_json = ada_part.split("window.ADA_DATA = ", 1)[1].rstrip().rstrip(";").strip()
    ada = json.loads(ada_json)
    return ada.get("itemNames", {}), ada.get("vpNames", {}), ada.get("gerNames", {})


def cola_actual(html):
    """Conserva tal cual los bloques finales window.ADA_DICT (cecos -> override de
    VP por CECO) y window.ADA_LOGO (logo del header), que no se derivan del Excel."""
    m = _asset_regex().search(html)
    if not m:
        raise ValueError("No encontré el bloque de datos (data.js) dentro del HTML.")
    data_js = gzip.decompress(base64.b64decode(m.group(2))).decode("utf-8")
    idx = data_js.find("\nwindow.ADA_DICT")
    if idx == -1:
        idx = data_js.find("window.ADA_DICT")
        return ("\n" + data_js[idx:]) if idx != -1 else ""
    return data_js[idx:]


def construir_data_js(records, vps, gers, items, maps, dist, cola=""):
    itemNames, vpNames, gerNames = maps
    ada = {"records": records, "vps": vps, "gers": gers, "items": items,
           "itemNames": itemNames, "vpNames": vpNames, "gerNames": gerNames}
    j = lambda o: json.dumps(o, ensure_ascii=False, separators=(",", ":"))
    return ("window.ADA_DATA = " + j(ada) + ";\n" +
            "window.DIST_DATA = " + j({"records": dist}) + ";" + cola)


def inyectar(html, data_js):
    new_b64 = base64.b64encode(gzip.compress(data_js.encode("utf-8"), mtime=0)).decode("ascii")
    nuevo, n = _asset_regex().subn(lambda m: m.group(1) + new_b64 + m.group(3), html)
    if n != 1:
        raise ValueError(f"Esperaba reemplazar 1 bloque de datos, encontré {n}.")
    return nuevo


# ------------------------------- proceso -------------------------------------
def actualizar(bbdd_path, dist_path, html_path, log):
    if not os.path.isfile(html_path):
        raise ValueError("No encuentro el HTML del dashboard:\n" + html_path)
    log("Leyendo Excel corporativo…");      records = leer_bbdd(bbdd_path, log)
    log("Leyendo Excel de distribuibles…"); dist = leer_distribuibles(dist_path, log)
    vps, gers, items = _uniq(r["vp"] for r in records), _uniq(r["ger"] for r in records), _uniq(r["item"] for r in records)
    log(f"  Únicos: {len(vps)} VP · {len(gers)} gerencias · {len(items)} ítems.")

    log("Leyendo el HTML actual…")
    with open(html_path, "r", encoding="utf-8", newline="") as f:
        html = f.read()
    maps = maps_actuales(html)
    cola = cola_actual(html)
    log(f"  Mapas de nombres conservados: {len(maps[0])} ítems · {len(maps[1])} VP · {len(maps[2])} gerencias.")
    log(f"  Bloques ADA_DICT/ADA_LOGO conservados: {len(cola)} chars.")

    data_js = construir_data_js(records, vps, gers, items, maps, dist, cola)
    # respaldo del HTML anterior
    shutil.copyfile(html_path, html_path + ".bak")
    log("Respaldo creado: " + os.path.basename(html_path) + ".bak")

    nuevo = inyectar(html, data_js)
    with open(html_path, "w", encoding="utf-8", newline="") as f:
        f.write(nuevo)
    # guardar también data.js suelto (referencia)
    with open(os.path.join(os.path.dirname(html_path), "data.js"), "w", encoding="utf-8", newline="") as f:
        f.write(data_js)
    log("HTML actualizado e inyectado.")

    # zip para enviar
    zip_path = os.path.join(os.path.dirname(html_path), "Dashboard Actividad Corporativa.zip")
    readme = ("DASHBOARD - Actividad Corporativa + Distribuibles (AMSA)\r\n\r\n"
              "Doble clic en 'Dashboard Actividad Corporativa.html' (Chrome/Edge/Firefox).\r\n"
              "No requiere instalar nada ni internet. Si sale en blanco: clic derecho -> Abrir con -> Chrome/Edge.\r\n"
              f"\r\nActualizado: {datetime.date.today().strftime('%d-%m-%Y')}\r\n")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        z.write(html_path, arcname=HTML_NAME)
        z.writestr("LEER - Como abrir.txt", readme)
    log("ZIP listo: " + os.path.basename(zip_path))
    return zip_path


# --------------------------------- GUI ---------------------------------------
def autodetectar():
    up = os.path.join(HERE, "uploads")
    bbdd = dist = ""
    if os.path.isdir(up):
        for f in os.listdir(up):
            fl = f.lower()
            if fl.endswith(".xlsx"):
                if not bbdd and ("histórico" in fl or "historico" in fl or "(original)" in fl):
                    bbdd = os.path.join(up, f)
                if not dist and "distribuible" in fl:
                    dist = os.path.join(up, f)
    return bbdd, dist


def main():
    root = tk.Tk()
    root.title("Actualizar Dashboard AMSA")
    root.geometry("720x520")
    root.configure(bg="#faf9f5")

    PAD = dict(padx=14, pady=6)
    tk.Label(root, text="Actualizar dashboard — Actividad Corporativa + Distribuibles",
             font=("Segoe UI", 13, "bold"), bg="#faf9f5", fg="#14515a").pack(anchor="w", **PAD)
    tk.Label(root, text="Elige los Excel actualizados y pulsa «Actualizar dashboard».",
             bg="#faf9f5", fg="#5B5C64").pack(anchor="w", padx=14)

    bbdd0, dist0 = autodetectar()
    html0 = os.path.join(HERE, HTML_NAME)

    def fila(label, valor):
        fr = tk.Frame(root, bg="#faf9f5"); fr.pack(fill="x", **PAD)
        tk.Label(fr, text=label, width=22, anchor="w", bg="#faf9f5").pack(side="left")
        var = tk.StringVar(value=valor)
        ent = tk.Entry(fr, textvariable=var); ent.pack(side="left", fill="x", expand=True, padx=6)
        def pick():
            p = filedialog.askopenfilename(title=label, filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")])
            if p: var.set(p)
        tk.Button(fr, text="Examinar…", command=pick).pack(side="left")
        return var

    v_bbdd = fila("Excel Corporativo (BBDD):", bbdd0)
    v_dist = fila("Excel Distribuibles:", dist0)
    v_html = fila("Dashboard HTML a actualizar:", html0)

    logbox = tk.Text(root, height=14, bg="#1f2428", fg="#e6e6e6",
                     insertbackground="#fff", font=("Consolas", 9), relief="flat")
    logbox.pack(fill="both", expand=True, padx=14, pady=(8, 4))

    def log(msg):
        logbox.insert("end", msg + "\n"); logbox.see("end"); root.update_idletasks()

    def run():
        logbox.delete("1.0", "end")
        for nombre, var in (("Excel Corporativo", v_bbdd), ("Excel Distribuibles", v_dist), ("HTML", v_html)):
            if not var.get() or not os.path.isfile(var.get()):
                messagebox.showerror("Falta archivo", f"Selecciona un archivo válido para: {nombre}.")
                return
        btn.config(state="disabled")
        try:
            zip_path = actualizar(v_bbdd.get(), v_dist.get(), v_html.get(), log)
            log("\n✓ LISTO. Dashboard actualizado.")
            if messagebox.askyesno("Listo", "Dashboard actualizado.\n\n¿Abrir la carpeta con el HTML y el ZIP?"):
                os.startfile(os.path.dirname(v_html.get()))
        except Exception as e:
            log("\n✗ ERROR: " + str(e))
            log(traceback.format_exc())
            messagebox.showerror("Error", str(e))
        finally:
            btn.config(state="normal")

    btn = tk.Button(root, text="Actualizar dashboard", command=run,
                    bg="#2a8a96", fg="white", font=("Segoe UI", 11, "bold"),
                    relief="flat", padx=16, pady=8, cursor="hand2")
    btn.pack(pady=(2, 12))

    root.mainloop()


if __name__ == "__main__":
    main()
