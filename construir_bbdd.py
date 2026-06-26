# -*- coding: utf-8 -*-
"""
Construir BBDD central (parquet) + regenerar el Dashboard
=========================================================
Centraliza TODA la información de gastos (Corporativo + Distribuible, todos los
años) en una única base de datos en formato **parquet** y, a partir de ella,
regenera el `data.js` que consume el HTML autocontenido.

  Excel  ──►  bbdd_corporativo.parquet  (única fuente de verdad)  ──►  data.js  ──►  HTML

Diseño 2026 (especial): el Excel corporativo trae los meses 2026.01–2026.05 y un
total anual 2026.TOTAL. Generamos DOS series de tiempo independientes:
  · 2026 YTD  → Real y Ppto acumulados de enero a mayo (suma de los 5 meses).
  · 2026 FY   → SOLO Ppto (presupuesto anual = 2026.TOTAL Plan). El Real del
                total NO se usa.

Requisitos: Python 3 + openpyxl + pandas + pyarrow.
Uso:  python construir_bbdd.py        (CLI, sin interfaz gráfica)

El parquet es la base de datos: tabla "tidy" (larga) con una fila por
  (src, comp, vp, ger, item, year, period)  y columnas  real / plan.
  · src    : 'corp' | 'dist'
  · comp   : None (corporativo) | 'MLP' | 'ANT' | 'CEN' | 'CMZ'
  · period : 'TOTAL' (anual) | '01'..'12' (mensual, hoy solo 2026.01–05)
"""
import os, json, re, datetime, shutil, zipfile, gzip, base64, sys
import openpyxl
import pandas as pd

# Reutilizamos los diccionarios y utilidades del actualizador existente
# (mapas VP↔Gerencia, normalización de celdas y los helpers que inyectan el
#  data.js dentro del HTML). Importarlo NO abre la interfaz gráfica.
import actualizar_dashboard as ad

HERE = os.path.dirname(os.path.abspath(__file__))
UP = os.path.join(HERE, "uploads")

# --- Archivos fuente -------------------------------------------------------
CORP_FILE = os.path.join(UP, "Gastos Act Corpor histórico 2026 FY.xlsx")
DIST_FULL = os.path.join(UP, "20260526 DISTRIBUIBLES_CIAS_2022_2025 full.xlsx")
DIST_2026 = os.path.join(UP, "20260526 DISTRIBUIBLES_CIAS_2026 FY.xlsx")
CECOS_FILE = os.path.join(UP, "CECOS.xlsx")   # diccionario por código (VP/Gerencia/Tipo Costo/¿Aplica?)
DOT_FILE = os.path.join(UP, "Dotaciones Histórico AMSA.xlsx")  # dotaciones (FTE) por VP/Gerencia
PARQUET_DOT = os.path.join(HERE, "dotaciones.parquet")

PARQUET = os.path.join(HERE, "bbdd_corporativo.parquet")
HTML_PATH = os.path.join(HERE, ad.HTML_NAME)

YEARS_HIST = [2022, 2023, 2024, 2025]   # años con solo total anual
MESES_2026 = ["01", "02", "03", "04", "05"]  # meses cargados de 2026 (YTD)
COMPS = ["MLP", "ANT", "CEN", "CMZ"]

# Normalización de nombres de Gerencia: SAP corta los nombres a ~20 caracteres y a
# veces el recorte no coincide con cómo está escrita la MISMA Gerencia en el
# Diccionario. Esta tabla equipara recorte → nombre del Diccionario. SOLO ajusta el
# nombre (identidad de la Gerencia); NO elige VP. Edita aquí si aparece un recorte
# nuevo que no cruza con el Diccionario.  Formato:  "recorte SAP": "nombre Diccionario".
GER_RENAME = {
    "Gcia Téc d Min y Pro":  "Gerencia Minería y Procesos",
    "Prog Compet Costos":    "Programa competitividad de costos",
    "Generalistas AMSA":     "Generalistas",
    "Gcia.Pr.nva form tra":  "Grcia Proy. Nuevas formas de trabajar",
    "Gcia TICA":             "TICA corporativo",
    "Transformación Digit":  "Transformación Digital",
}

# Gerencias que NO están en el Diccionario por nombre pero que model.js ya resuelve
# vía GER_ALIAS (unificación de recortes SAP → VP). No se marcan como "sin VP".
ALIAS_HANDLED = {
    "Gcia.Riesgo y Ctrl I",                          # → Grcia.Riesg.CompyCIn / VP Finanzas
    "Grcia. Riesgos, Compliance y Control Interno",
}

# Marcador para gerencias sin código/VP que se dejan en la pestaña Diccionario.
SINCOD = "SINCOD:"   # prefijo de "CECO" placeholder · VP vacía = pendiente de asignar


def log(msg):
    # Consola Windows (cp1252) no soporta algunos símbolos; degradar sin romper.
    try:
        print(msg, flush=True)
    except UnicodeEncodeError:
        enc = sys.stdout.encoding or "cp1252"
        print(msg.encode(enc, "replace").decode(enc), flush=True)


# ===========================================================================
#  1) LECTURA DE EXCEL  →  filas "tidy"
# ===========================================================================
def _dic_cecos(cola):
    """Lee la lista CORP_DICT.cecos (CECO → Gerencia → VP) del bloque conservado."""
    i = cola.find("{cecos:")
    if i == -1:
        return []
    j = cola.find("[", i)
    try:
        arr, _ = json.JSONDecoder().raw_decode(cola[j:])
        return arr
    except ValueError:
        return []


def actualizar_diccionario(cola, cecos, vpNames):
    """Actualiza la pestaña Diccionario (CORP_DICT) contra CECOS.xlsx:
      · sincroniza el VP de los códigos existentes (si el VP real difiere),
      · AGREGA los códigos que faltan (p. ej. los distribuibles 1001AD…),
        con su Gerencia (Desc. CECO) y VP.
    El VP se guarda con el nombre largo (consistente con dispVP del tablero)."""
    tag = "window.CORP_DICT={cecos:"
    i = cola.find(tag)
    if i == -1:
        return cola, (0, 0)
    j = i + len(tag)
    arr, end = json.JSONDecoder().raw_decode(cola[j:])
    inv = {v: k for k, v in (vpNames or {}).items()}
    short_of = lambda v: ad.VP_DICT.get(v) or inv.get(v) or v
    present = {ad._txt(c.get("c")) for c in arr}
    nsync = 0
    for c in arr:
        code = ad._txt(c.get("c"))
        if code in cecos:
            vp, ger, tc, ap = cecos[code]
            c["tc"] = tc or None     # Tipo Costo (C1/C3/Comercialización)
            c["ap"] = ap or None     # ¿Aplica? (Sí/No)
            if short_of(ad._txt(c.get("v"))) != vp:
                c["v"] = vpNames.get(vp, vp); nsync += 1
    nadd = 0
    for code in sorted(cecos):
        if code not in present:
            vp, ger, tc, ap = cecos[code]
            arr.append({"c": code, "g": ger, "v": vpNames.get(vp, vp), "tc": tc or None, "ap": ap or None}); nadd += 1
    nueva = (cola[:j] + json.dumps(arr, ensure_ascii=False, separators=(",", ":"))
             + cola[j + end:])
    return nueva, (nsync, nadd)


def leer_corp(code2vp=None, cecos=None):
    """Corporativo: col A = Código CECO (relleno), col B = Gerencia (relleno),
    col C = Ítem. Años en cols (real, plan): 2022→3/4, 2023→6/7, 2024→9/10,
    2025→12/13.  2026 mensual: 05→15/16, 04→18/19, 03→21/22, 02→24/25, 01→27/28.
    2026 total anual: 30 (real) / 31 (plan).
    La VP se asigna por el CÓDIGO de la Gerencia contra el Diccionario (cruce directo);
    Tipo Costo y ¿Aplica? también vienen por código desde CECOS.xlsx."""
    wb = openpyxl.load_workbook(CORP_FILE, data_only=True, read_only=True)
    sheet = next((s for s in wb.sheetnames if s.strip() == ad.SHEET_CORP), None)
    if sheet is None:
        raise ValueError(f"El Excel corporativo no tiene la hoja '{ad.SHEET_CORP}'.")
    ws = wb[sheet]
    ANUAL = {2022: 3, 2023: 6, 2024: 9, 2025: 12}     # col del Real (Plan = +1)
    MES = {"05": 15, "04": 18, "03": 21, "02": 24, "01": 27}
    TOT26 = 30
    dic = code2vp or {}
    cec = cecos or {}
    rows, code, ger, orphans = [], None, None, {}
    for r in ws.iter_rows(min_row=10, values_only=True):
        c0, c1, c2 = r[0], r[1], r[2]
        if c0 not in (None, ""):
            code = ad._txt(c0)
        if c1 not in (None, ""):
            # GER_RENAME unifica recortes SAP al nombre que usan los mapas de nombres.
            ger = GER_RENAME.get(ad._txt(c1), ad._txt(c1))
        if c2 in (None, ""):
            continue
        # VP por CÓDIGO contra el Diccionario (cruce directo). Sin código o código
        # ausente del Diccionario → queda como su propia VP y se MARCA en el Diccionario.
        vp = dic.get(code)
        if vp is None:
            orphans[ger] = code or ""; vp = ger
        # Tipo Costo y ¿Aplica? por código desde CECOS (mismo cruce).
        _, _, tc, ap = cec.get(code, ("", "", "", "Sí"))
        item = ad._txt(c2)
        g = lambda i: ad._num(r[i]) if i < len(r) else None
        base = dict(src="corp", comp=None, vp=vp, ger=ger, item=item, tipo_costo=tc, aplica=ap)
        for y, col in ANUAL.items():
            rows.append(dict(base, year=y, period="TOTAL", real=g(col), plan=g(col + 1)))
        for mm, col in MES.items():
            rows.append(dict(base, year=2026, period=mm, real=g(col), plan=g(col + 1)))
        rows.append(dict(base, year=2026, period="TOTAL", real=g(TOT26), plan=g(TOT26 + 1)))
    wb.close()
    log(f"  Corporativo: {len(rows)} filas tidy.")
    return rows, orphans


def cecos_por_codigo(vpNames):
    """Diccionario por CÓDIGO CECO desde CECOS.xlsx, hoja 'CECOS Corporativo':
    código → (VP clave corta, Gerencia, Tipo Costo, ¿Aplica?). col 'CECO' = código,
    'Desc. Nodo 2 (VP)' = VP, 'Desc. CECO' = Gerencia, 'Tipo Costo' y '¿Aplica?' tal
    cual. Cubre AC (corporativo) y AD (distribuible). VP largo → clave corta con
    VP_DICT + inverso vpNames. ¿Aplica? se normaliza a 'Sí'/'No'."""
    inv = {v: k for k, v in (vpNames or {}).items()}
    wb = openpyxl.load_workbook(CECOS_FILE, data_only=True, read_only=True)
    sheet = next((s for s in wb.sheetnames if "orporativo" in s), wb.sheetnames[-1])
    grid = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    wb.close()
    h = grid[0]
    idx = {ad._txt(c): i for i, c in enumerate(h)}
    cC = idx.get("CECO")
    gerC = idx.get("Desc. CECO")
    vpC = next((i for i, c in enumerate(h) if "Nodo 2" in ad._txt(c) and "Desc" in ad._txt(c)), None)  # 'Desc. Nodo 2 (VP)' (no el código 'Nodo 2')
    tcC = next((i for i, c in enumerate(h) if "Tipo Costo" in ad._txt(c)), None)
    apC = next((i for i, c in enumerate(h) if "Aplica" in ad._txt(c)), None)
    out = {}
    for r in grid[1:]:
        code = ad._txt(r[cC]) if cC is not None and cC < len(r) else ""
        if not code:
            continue
        vp_raw = ad._txt(r[vpC]) if vpC is not None and vpC < len(r) else ""
        ger = ad._txt(r[gerC]) if gerC is not None and gerC < len(r) else ""
        tc = ad._txt(r[tcC]) if tcC is not None and tcC < len(r) else ""
        ap = "No" if (ad._txt(r[apC]) if apC is not None and apC < len(r) else "") == "No" else "Sí"
        out[code] = (ad.VP_DICT.get(vp_raw) or inv.get(vp_raw) or vp_raw, ger, tc, ap)
    return out


def leer_dist(dist_dic=None):
    """Distribuible: valores 2022-2025 de la planilla 'full' y 2026 (meses ene–may +
    total anual) de la planilla 'FY'. La VP y la Gerencia se asignan por CÓDIGO CECO
    contra CECOS.xlsx (cruce directo); el Ítem viene de cada planilla."""
    dic = dist_dic or {}
    sin_dic = set()

    def vg(code):
        v = dic.get(code)
        if v is None:
            sin_dic.add(code)
            return (code, code, "", "Sí")
        return v

    rows = []
    # ---- 2022-2025 (full): código en 'CECO Código', ítem en 'Item Relevante' ----
    wb = openpyxl.load_workbook(DIST_FULL, data_only=True, read_only=True)
    for cm in COMPS:
        grid = [list(r) for r in wb[cm].iter_rows(values_only=True)]
        hy = next((i for i in range(min(15, len(grid)))
                   if any(str(c) == "2022.TOTAL" for c in grid[i])), None)
        if hy is None:
            raise ValueError(f"No encontré los años en {cm} (full).")
        yc = {y: ci for ci, c in enumerate(grid[hy]) for y in YEARS_HIST
              if str(c) == f"{y}.TOTAL"}
        hl = grid[hy + 1]
        idx = {ad._txt(c): i for i, c in enumerate(hl)}
        itC = idx.get("Item Relevante")
        codeC = next((i for i, c in enumerate(hl) if "digo" in ad._txt(c)), None)
        for ri in range(hy + 2, len(grid)):
            row = grid[ri]
            it = ad._txt(row[itC]) if itC is not None and itC < len(row) else ""
            code = ad._txt(row[codeC]) if codeC is not None and codeC < len(row) else ""
            if not it or not code:
                continue
            vp, ger, tc, ap = vg(code)
            g = lambda i: ad._num(row[i]) if i is not None and i < len(row) else None
            for y in YEARS_HIST:
                rows.append(dict(src="dist", comp=cm, vp=vp, ger=ger, item=it,
                                 tipo_costo=tc, aplica=ap, year=y, period="TOTAL",
                                 real=g(yc[y]), plan=g(yc[y] + 1)))
    wb.close()

    # ---- 2026 (FY): col0 = 'AMSA<código> - <ceco>', col1 = '<x> - <item>' ----
    wb = openpyxl.load_workbook(DIST_2026, data_only=True, read_only=True)
    pat0 = re.compile(r'^AMSA(\S+)\s*-\s*(.*)$')
    pat1 = re.compile(r'^\S+\s*-\s*(.*)$')
    for cm in COMPS:
        grid = [list(r) for r in wb[cm].iter_rows(values_only=True)]
        hy = next((i for i in range(min(15, len(grid)))
                   if any(str(c) == "2026.01" for c in grid[i])), None)
        if hy is None:
            raise ValueError(f"No encontré 2026.01 en {cm} (FY).")
        pc = {}  # periodo → col del Real (Plan = +1)
        for ci, c in enumerate(grid[hy]):
            s = str(c)
            if s.startswith("2026.") and (s[5:] in MESES_2026 or s == "2026.TOTAL"):
                pc["TOTAL" if s == "2026.TOTAL" else s[5:]] = ci
        for ri in range(hy + 2, len(grid)):
            row = grid[ri]
            c0 = ad._txt(row[0]) if 0 < len(row) else ""
            m0 = pat0.match(c0) if c0 else None
            if not m0:
                continue
            code = m0.group(1)
            c1 = ad._txt(row[1]) if 1 < len(row) else ""
            m1 = pat1.match(c1)
            item = (m1.group(1) if m1 else c1).strip()
            vp, ger, tc, ap = vg(code)
            g = lambda i: ad._num(row[i]) if i is not None and i < len(row) else None
            for per, ci in pc.items():
                rows.append(dict(src="dist", comp=cm, vp=vp, ger=ger, item=item,
                                 tipo_costo=tc, aplica=ap,
                                 year=2026, period=per, real=g(ci), plan=g(ci + 1)))
    wb.close()
    if sin_dic:
        log(f"  AVISO dist: {len(sin_dic)} códigos sin entrada en CECOS.xlsx (se usan tal cual): "
            + ", ".join(sorted(sin_dic)[:8]) + (" …" if len(sin_dic) > 8 else ""))
    log(f"  Distribuible: {len(rows)} filas tidy.")
    return rows


# ===========================================================================
#  1b) DOTACIONES (FTE) — hojas "Consolidado Propios" / "Consolidado Contratista"
# ===========================================================================
def leer_dotaciones():
    """Lee las hojas consolidadas de Dotaciones (FTE, promedio anual Nº) por VP /
    Gerencia. Estructura: col0 = VP (sin sangría) / Gerencia (con sangría); años en
    cols Real/Ppto: 2022→1/2, 2023→3/4, 2024→5/6, 2025→7/8, 2026 YTD→9/10, 2026 FY→11.
    Devuelve (records para el dashboard, filas tidy para el parquet)."""
    if not os.path.isfile(DOT_FILE):
        log("  AVISO: no encontré el Excel de Dotaciones; se omite.")
        return [], []
    ANUAL = {2022: 1, 2023: 3, 2024: 5, 2025: 7}     # col del Real (Ppto = +1)
    SHEETS = {"Consolidado Propios": "propios", "Consolidado Contratista": "contratista"}
    wb = openpyxl.load_workbook(DOT_FILE, data_only=True, read_only=True)
    records, tidy = [], []
    for sheet, src in SHEETS.items():
        if sheet not in wb.sheetnames:
            log(f"  AVISO Dotaciones: falta la hoja '{sheet}'.")
            continue
        grid = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
        vp = None
        for r in grid[5:]:                       # datos desde la fila 6
            c0 = r[0] if r else None
            if c0 in (None, ""):
                continue
            s = str(c0)
            label = s.strip()
            if not (s.startswith(" ") or s.startswith("\t")):
                vp = label                       # fila VP (sin sangría)
                continue
            ger = label                          # fila Gerencia (con sangría) = registro
            g = lambda i: ad._num(r[i]) if i < len(r) else None
            rec = {"src": src, "vp": vp, "ger": ger}
            for y, col in ANUAL.items():
                rec[f"y{y}"] = {"real": g(col), "plan": g(col + 1)}
                tidy.append(dict(src=src, vp=vp, ger=ger, year=y, period="TOTAL",
                                 real=g(col), plan=g(col + 1)))
            rec["y2026"] = {"real": g(9), "plan": g(10)}       # YTD ene–may (real y ppto)
            rec["y2026fy"] = {"plan": g(11)}                   # presupuesto anual (solo ppto)
            tidy.append(dict(src=src, vp=vp, ger=ger, year=2026, period="YTD", real=g(9), plan=g(10)))
            tidy.append(dict(src=src, vp=vp, ger=ger, year=2026, period="FY", real=None, plan=g(11)))
            records.append(rec)
    wb.close()
    log(f"  Dotaciones: {len(records)} registros (Propios + Contratista).")
    return records, tidy


# ===========================================================================
#  2) PARQUET (base de datos central)
# ===========================================================================
def construir_parquet(code2vp=None, cecos=None):
    log("Leyendo Excel corporativo…");      corp, orphans = leer_corp(code2vp, cecos)
    log("Leyendo Excel de distribuibles…"); dist = leer_dist(cecos)
    df = pd.DataFrame(corp + dist,
                      columns=["src", "comp", "vp", "ger", "item",
                               "tipo_costo", "aplica", "year", "period", "real", "plan"])
    # Normaliza al grano (src,comp,vp,ger,item,tipo_costo,aplica,year,period) sumando
    # duplicados (la planilla full trae filas repetidas por mismo VP/Gerencia/Ítem).
    df = (df.groupby(["src", "comp", "vp", "ger", "item", "tipo_costo", "aplica", "year", "period"],
                     dropna=False, as_index=False)[["real", "plan"]].sum(min_count=1))
    df.to_parquet(PARQUET, index=False)
    log(f"Base de datos escrita: {os.path.basename(PARQUET)}  ({len(df)} filas)")
    return df, orphans


def marcar_en_diccionario(cola, orphans):
    """Deja en CORP_DICT.cecos una fila MARCADA (VP vacía = pendiente) por cada
    Gerencia que el Diccionario no resuelve, para asignarle ahí la VP. Usa el código
    real si lo hay; si no, un placeholder SINCOD. Primero limpia los marcadores
    previos del pipeline (idempotente): si una Gerencia ya quedó resuelta, su marca
    desaparece. `orphans` = {gerencia: código (o '')}."""
    tag = "window.CORP_DICT={cecos:"
    i = cola.find(tag)
    if i == -1:
        if orphans:
            log("  AVISO: no encontré CORP_DICT; no se pudo marcar en el Diccionario.")
        return cola, []
    j = i + len(tag)
    arr, end = json.JSONDecoder().raw_decode(cola[j:])
    # 1) limpia marcadores previos del pipeline: filas con VP vacía y código placeholder
    #    o cuya Gerencia ya quedó resuelta (ya no es huérfana).
    arr = [c for c in arr if not (ad._txt(c.get("v")) == "" and
           (str(c.get("c", "")).startswith(SINCOD) or ad._txt(c.get("g")) in orphans))]
    presentes = {ad._txt(c.get("g")) for c in arr}
    marcados = []
    for ger in sorted(orphans):
        if ger in ALIAS_HANDLED or ger in presentes:
            continue
        code = orphans[ger]
        arr.append({"c": code or (SINCOD + ger), "g": ger, "v": ""})   # VP vacía = pendiente
        marcados.append(f"{ger} ({code or 'sin código'})")
    nueva = (cola[:j] + json.dumps(arr, ensure_ascii=False, separators=(",", ":"))
             + cola[j + end:])
    return nueva, marcados


# ===========================================================================
#  3) PARQUET  →  data.js  (registros que consume el HTML)
# ===========================================================================
def _round(v):
    if v is None or pd.isna(v):
        return None
    f = round(float(v), 2)
    return int(f) if f == int(f) else f


def registros_desde_parquet(df, src):
    """Pivota la BBDD a la estructura de registro del dashboard, conservando el
    orden de aparición (id estable)."""
    sub = df[df["src"] == src]
    recs, seen = [], {}
    # índice rápido por clave → {(year,period): (real,plan)}
    keys_order = []
    bucket = {}
    for row in sub.itertuples(index=False):
        key = (row.comp, row.vp, row.ger, row.item, row.tipo_costo, row.aplica)
        if key not in bucket:
            bucket[key] = {}; keys_order.append(key)
        bucket[key][(row.year, row.period)] = (row.real, row.plan)
    for key in keys_order:
        comp, vp, ger, item, tc, ap = key
        cells = bucket[key]
        rec = {}
        if src == "dist":
            rec["comp"] = comp
        rec.update(vp=vp, ger=ger, item=item)
        # Clasificación por código (CECOS): Tipo Costo y ¿Aplica?. '' → None.
        rec["tc"] = (tc or None) if not (tc is None or pd.isna(tc)) else None
        rec["ap"] = ap if not (ap is None or pd.isna(ap)) else None
        for y in YEARS_HIST:
            r, p = cells.get((y, "TOTAL"), (None, None))
            rec[f"y{y}"] = {"real": _round(r), "plan": _round(p)}
        # 2026 YTD = suma ene–may (real y ppto). None/NaN se tratan como 0.
        nz = lambda v: 0 if (v is None or pd.isna(v)) else v
        ytd_r = sum(nz(cells.get((2026, mm), (0, 0))[0]) for mm in MESES_2026)
        ytd_p = sum(nz(cells.get((2026, mm), (0, 0))[1]) for mm in MESES_2026)
        rec["y2026"] = {"real": _round(ytd_r), "plan": _round(ytd_p)}
        # 2026 FY = SOLO ppto anual (2026.TOTAL Plan). Real no se usa.
        _, fy_p = cells.get((2026, "TOTAL"), (None, None))
        rec["y2026fy"] = {"plan": _round(fy_p)}
        recs.append(rec)
    return recs


def construir_data_js(df, maps, cola, dot_records=None):
    itemNames, vpNames, gerNames = maps
    corp = registros_desde_parquet(df, "corp")
    dist = registros_desde_parquet(df, "dist")
    vps = ad._uniq(r["vp"] for r in corp)
    gers = ad._uniq(r["ger"] for r in corp)
    items = ad._uniq(r["item"] for r in corp)
    ada = {"records": corp, "vps": vps, "gers": gers, "items": items,
           "itemNames": itemNames, "vpNames": vpNames, "gerNames": gerNames}
    j = lambda o: json.dumps(o, ensure_ascii=False, separators=(",", ":"))
    return ("window.CORP_DATA = " + j(ada) + ";\n" +
            "window.DIST_DATA = " + j({"records": dist}) + ";\n" +
            "window.DOT_DATA = " + j({"records": dot_records or []}) + ";" + cola), corp, dist


# ===========================================================================
#  4) Validación rápida contra el data.js vigente (no bloquea, informa)
# ===========================================================================
def validar(df):
    tot = (df[df.year.isin(YEARS_HIST) & (df.period == "TOTAL")]
           .groupby("year")[["real", "plan"]].sum())
    log("  Totales 2022-2025 (MM USD)  real / ppto:")
    for y in YEARS_HIST:
        r, p = tot.loc[y, "real"], tot.loc[y, "plan"]
        log(f"    {y}: {r/1e6:,.1f} / {p/1e6:,.1f}")
    d26 = df[df.year == 2026]
    ytd_r = d26[d26.period.isin(MESES_2026)]["real"].sum()
    ytd_p = d26[d26.period.isin(MESES_2026)]["plan"].sum()
    fy_p = d26[d26.period == "TOTAL"]["plan"].sum()
    log(f"  2026 YTD (ene-may)  real / ppto:  {ytd_r/1e6:,.1f} / {ytd_p/1e6:,.1f} MM USD")
    log(f"  2026 FY  (ppto anual):            {fy_p/1e6:,.1f} MM USD")


# ===========================================================================
#  5) Regenerar el HTML (inyectar data.js) + zip
# ===========================================================================
def regenerar_html(data_js):
    if not os.path.isfile(HTML_PATH):
        raise ValueError("No encuentro el HTML del dashboard:\n" + HTML_PATH)
    with open(HTML_PATH, "r", encoding="utf-8", newline="") as f:
        html = f.read()
    shutil.copyfile(HTML_PATH, HTML_PATH + ".bak")
    nuevo = ad.inyectar(html, data_js)
    with open(HTML_PATH, "w", encoding="utf-8", newline="") as f:
        f.write(nuevo)
    with open(os.path.join(HERE, "data.js"), "w", encoding="utf-8", newline="") as f:
        f.write(data_js)
    log("HTML actualizado e inyectado · data.js regenerado.")
    zip_path = os.path.join(HERE, "Dashboard Actividad Corporativa.zip")
    readme = ("DASHBOARD - Actividad Corporativa + Distribuibles (AMSA)\r\n\r\n"
              "Doble clic en 'Dashboard Actividad Corporativa.html' (Chrome/Edge/Firefox).\r\n"
              f"\r\nActualizado: {datetime.date.today().strftime('%d-%m-%Y')}\r\n")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        z.write(HTML_PATH, arcname=ad.HTML_NAME)
        z.writestr("LEER - Como abrir.txt", readme)
    log("ZIP listo: " + os.path.basename(zip_path))


# ===========================================================================
def main():
    log("=== Construir BBDD central (parquet) + regenerar dashboard ===")
    log("Leyendo el HTML actual (mapas de nombres y Diccionario conservados)…")
    with open(HTML_PATH, "r", encoding="utf-8", newline="") as f:
        html = f.read()
    maps = ad.maps_actuales(html)            # (itemNames, vpNames, gerNames)
    cola = ad.cola_actual(html)              # window.CORP_DICT (CECO→Gerencia→VP) + CORP_LOGO
    cecos = cecos_por_codigo(maps[1])        # ÚNICA fuente: CECOS.xlsx (VP/Gerencia/Tipo Costo/¿Aplica?)
    log(f"  CECOS.xlsx: {len(cecos)} códigos.")
    # Unifica TODO a CECOS: actualiza la pestaña Diccionario (sincroniza VP + agrega
    # los códigos que falten, p. ej. distribuibles) y mapea corp por código.
    cola, (nsync, nadd) = actualizar_diccionario(cola, cecos, maps[1])
    log(f"  Diccionario (pestaña): {nsync} VP sincronizadas · {nadd} códigos agregados.")
    dic = {code: vg[0] for code, vg in cecos.items()}   # código → VP corta (CECOS)
    df, orphans = construir_parquet(dic, cecos)
    validar(df)
    # Gerencias que el Diccionario no resuelve → se MARCAN en la pestaña Diccionario.
    cola, marcadas = marcar_en_diccionario(cola, orphans)
    if marcadas:
        log("  Marcadas en el Diccionario (asignar VP ahí): " + ", ".join(marcadas))
    aliased = sorted(g for g in orphans if g in ALIAS_HANDLED)
    if aliased:
        log("  Resueltas por model.js (GER_ALIAS), no se marcan: " + ", ".join(aliased))
    # Dotaciones (FTE): records para el dashboard + parquet aparte.
    log("Leyendo Dotaciones (FTE)…")
    dot_records, dot_tidy = leer_dotaciones()
    if dot_tidy:
        pd.DataFrame(dot_tidy, columns=["src", "vp", "ger", "year", "period", "real", "plan"]) \
            .to_parquet(PARQUET_DOT, index=False)
        log(f"  Dotaciones escritas: {os.path.basename(PARQUET_DOT)} ({len(dot_tidy)} filas)")
    data_js, corp, dist = construir_data_js(df, maps, cola, dot_records)
    log(f"  Registros: {len(corp)} corp · {len(dist)} dist · {len(dot_records)} dotaciones.")
    regenerar_html(data_js)
    log("\n✓ LISTO.")


if __name__ == "__main__":
    main()
